"""
Streamlit GUI to aggregate and visualize tagged sentiment data from BestBuy review CSVs.

2025-06-18 highlights 🆕
------------------------------------------------
1. タブ切り替えで各モデル＆全体集計を閲覧
2. タグ別感情割合のモデル比較＋Excel「Tag_Ratios」シート
3. 日本語タイトルが文字化けしないようフォント自動設定
4. ★モデルごとの総レビュー件数をテーブル＆グラフに表示
5. ★件数ラベルを“棒の最上端（≒100%）”に配置
"""

# ---------------------------------------------------------------------------
# 日本語フォント自動セットアップ (matplotlib)
# ---------------------------------------------------------------------------
import matplotlib as mpl
import matplotlib.font_manager as fm

_FONT_CANDS = ["IPAexGothic", "Noto Sans CJK JP", "Yu Gothic", "MS Gothic"]


def _pick_jp_font() -> str | None:
    avail = {f.name for f in fm.fontManager.ttflist}
    for cand in _FONT_CANDS:
        if cand in avail:
            return cand
    return None


_SELECTED_FONT = _pick_jp_font()
if _SELECTED_FONT:
    mpl.rcParams["font.family"] = _SELECTED_FONT
    mpl.rcParams["axes.unicode_minus"] = False  # − を正しく表示

# ---------------------------------------------------------------------------
# 標準ライブラリ & サードパーティ
# ---------------------------------------------------------------------------
import io
from datetime import datetime
from typing import Dict, List
from uuid import uuid4

import matplotlib.pyplot as plt
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# ---------------------------------------------------------------------------
# Excel-Injection 回避ユーティリティ
# ---------------------------------------------------------------------------
def _escape_excel_formula(val):
    if isinstance(val, str) and val and val[0] in ("=", "+", "-", "@"):
        return "'" + val
    return val


def _safe_df(df: pd.DataFrame) -> pd.DataFrame:
    """DataFrame を Excel に書く前に安全化"""
    return df.applymap(_escape_excel_formula)


# ---------------------------------------------------------------------------
# Streamlit ページ設定
# ---------------------------------------------------------------------------
st.set_page_config(page_title="BestBuy Review Analyzer", layout="centered")

st.title("📊 BestBuy Review Analyzer")
st.markdown(
    "CSV ファイルを複数選択すると、タグ別感情・評価分布などを自動集計します。<br>"
    "モデル名はファイル名から最大 5 トークンまで抽出して表示します！",
    unsafe_allow_html=True,
)

if not _SELECTED_FONT:
    st.warning(
        "日本語フォントが見つかりません。グラフが文字化けする場合は "
        "`sudo apt-get install fonts-noto-cjk` などで追加してください。"
    )

# ---------------------------------------------------------------------------
# ファイルアップローダ
# ---------------------------------------------------------------------------
uploaded_files = st.file_uploader(
    "解析したい CSV を選択 (複数可)", type="csv", accept_multiple_files=True
)
if not uploaded_files:
    st.info("CSV をアップロードしてください。")
    st.stop()

# ---------------------------------------------------------------------------
# モデル名抽出ヘルパ
# ---------------------------------------------------------------------------
MAX_TOKENS = 5
STOPWORDS = {
    "eval",
    "evaluation",
    "result",
    "results",
    "analysis",
    "analyze",
    "output",
}


def derive_model_name(filename: str) -> str:
    stem = filename.rsplit("/", 1)[-1].rsplit(".", 1)[0].replace("-", "_")
    tokens = [t for t in stem.split("_") if t and t.lower() not in STOPWORDS]
    short = "_".join(tokens[:MAX_TOKENS]) or stem
    return short[:30]  # Excel シート名 31 文字制限


# ---------------------------------------------------------------------------
# ファイル読込 & 前処理
# ---------------------------------------------------------------------------
TAG_COLUMNS: List[str] = [
    "SoundQuality",
    "Music",
    "Movies",
    "Surround",
    "Dialogue",
    "Bass",
    "App",
    "Setup",
    "Design",
]
SENTIMENTS = ["Positive", "Neutral", "Negative"]

file_dfs: Dict[str, pd.DataFrame] = {}
model_names: Dict[str, str] = {}
dup_counter: Dict[str, int] = {}

for uf in uploaded_files:
    df = pd.read_csv(uf)
    df.columns = [c.strip() for c in df.columns]
    df["__source_file"] = uf.name
    file_dfs[uf.name] = df

    base = derive_model_name(uf.name)
    dup_counter[base] = dup_counter.get(base, 0) + 1
    model_names[uf.name] = base if dup_counter[base] == 1 else f"{base}-{dup_counter[base]}"

    uf.seek(0)  # Excel 書き出し用にポインタ巻き戻し

all_data = pd.concat(file_dfs.values(), ignore_index=True)
available_tags = [c for c in TAG_COLUMNS if c in all_data.columns]

# ---------------------------------------------------------------------------
# 集計関数
# ---------------------------------------------------------------------------
def build_tag_summary(df: pd.DataFrame) -> pd.DataFrame:
    return pd.DataFrame(
        {
            tag: df[tag].value_counts().reindex(SENTIMENTS, fill_value=0)
            for tag in available_tags
        }
    ).T


def calc_ratio_df(files: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = {}
    for fname, df in files.items():
        ratios = {}
        for tag in available_tags:
            cnt = df[tag].value_counts().reindex(SENTIMENTS, fill_value=0)
            total = int(cnt.sum()) or 1
            for s in SENTIMENTS:
                ratios[f"{tag}_{s}"] = round(cnt[s] / total * 100, 2)
            ratios[f"{tag}_Reviews"] = int(cnt.sum())
        rows[model_names[fname]] = ratios
    return pd.DataFrame.from_dict(rows, orient="index")


def build_tag_ratio_long(files: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    records = []
    for fname, df in files.items():
        model = model_names[fname]
        for tag in available_tags:
            cnt = df[tag].value_counts().reindex(SENTIMENTS, fill_value=0)
            total = int(cnt.sum()) or 1
            for s in SENTIMENTS:
                records.append(
                    {
                        "Model": model,
                        "Tag": tag,
                        "Sentiment": s,
                        "Ratio(%)": round(cnt[s] / total * 100, 2),
                        "Reviews": int(cnt[s]),  # 極性別件数
                    }
                )
    return pd.DataFrame(records)


def build_tag_ratio_long_all(files, all_data) -> pd.DataFrame:
    df_long = build_tag_ratio_long(files)
    for tag in available_tags:
        cnt = all_data[tag].value_counts().reindex(SENTIMENTS, fill_value=0)
        total = int(cnt.sum()) or 1
        for s in SENTIMENTS:
            df_long.loc[len(df_long)] = {
                "Model": "All_Files",
                "Tag": tag,
                "Sentiment": s,
                "Ratio(%)": round(cnt[s] / total * 100, 2),
                "Reviews": int(cnt[s]),
            }
    return df_long


@st.cache_data(show_spinner=False)
def cached_ratio_df(files):
    return calc_ratio_df(files)


@st.cache_data(show_spinner=False)
def cached_long_df(files, all_data):
    return build_tag_ratio_long_all(files, all_data)


ratio_df = cached_ratio_df(file_dfs)
tag_ratio_long_df = cached_long_df(file_dfs, all_data)

# ---------------------------------------------------------------------------
# 個別解析タブ
# ---------------------------------------------------------------------------
def render_single(df: pd.DataFrame):
    tbl = build_tag_summary(df)
    st.subheader("タグ別 ポジ/ネガ/ニュートラル 件数")
    with st.expander("表を表示", expanded=False):
        st.dataframe(tbl)

    fig, ax = plt.subplots()
    tbl.plot(kind="bar", stacked=True, ax=ax)
    ax.set_xlabel("Tag")
    ax.set_ylabel("Count")
    ax.legend(title="Sentiment")
    st.pyplot(fig)


tabs = st.tabs(
    (["All Files"] if len(file_dfs) > 1 else []) + [model_names[f] for f in file_dfs]
)

if len(file_dfs) > 1:
    with tabs[0]:
        st.header("【All Files (Aggregate)】")
        render_single(all_data)

start_idx = 0 if len(file_dfs) == 1 else 1
for tab, (fname, df) in zip(tabs[start_idx:], file_dfs.items()):
    with tab:
        st.header(f"【{model_names[fname]}】の解析結果")
        render_single(df)

# ---------------------------------------------------------------------------
# モデル比較グラフ
# ---------------------------------------------------------------------------
st.header("モデル比較: タグ別スコア割合 (Pos/Neu/Neg) + 件数")
chosen_tag = st.selectbox("比較したいタグ", available_tags)

view = tag_ratio_long_df[tag_ratio_long_df["Tag"] == chosen_tag]
pivot_view = view.pivot(index="Model", columns="Sentiment", values="Ratio(%)")
reviews_map = view.groupby("Model")["Reviews"].sum()  # ★合計に変更
pivot_view["Reviews"] = pivot_view.index.map(reviews_map).astype(int)
pivot_view = pivot_view[SENTIMENTS + ["Reviews"]]

with st.expander("比較表を表示", expanded=False):
    st.dataframe(pivot_view)

fig_cmp, ax_cmp = plt.subplots()
pivot_view[SENTIMENTS].plot(kind="bar", stacked=True, ax=ax_cmp)
ax_cmp.set_ylabel("Percentage (%)")
ax_cmp.set_title(chosen_tag)
ax_cmp.legend(title="Sentiment", bbox_to_anchor=(1.05, 1), loc="upper left")

for rect, total in zip(ax_cmp.containers[0], pivot_view["Reviews"]):
    x_center = rect.get_x() + rect.get_width() / 2
    ax_cmp.text(x_center, 101, f"{int(total)}", ha="center", va="bottom", fontsize=8)
ax_cmp.set_ylim(0, 105)

st.pyplot(fig_cmp)

# ---------------------------------------------------------------------------
# キーワード検索
# ---------------------------------------------------------------------------
st.header("キーワード検索 / フィルタ (All Files)")
kw = st.text_input("含めたいキーワード (スペース区切り可)")
if kw:
    mask = (
        all_data["review_text"].astype(str).str.contains(kw, case=False, na=False)
        | all_data.get("translated_text", pd.Series("", index=all_data.index))
        .astype(str)
        .str.contains(kw, case=False, na=False)
    )
    filtered = all_data[mask]
    st.write(f"該当レビュー数: {len(filtered)} 件")
    st.dataframe(filtered)
else:
    filtered = all_data

# ---------------------------------------------------------------------------
# Excel エクスポート
# ---------------------------------------------------------------------------
st.header("Excel エクスポート")

buf = io.BytesIO()
with pd.ExcelWriter(buf, engine="openpyxl") as writer:
    # メインシート
    _safe_df(build_tag_summary(all_data)).to_excel(writer, sheet_name="Tag_Summary_All")
    _safe_df(ratio_df).to_excel(writer, sheet_name="Tag_Ratios")
    _safe_df(tag_ratio_long_df).to_excel(writer, sheet_name="Tag_Ratios_Long", index=False)
    _safe_df(filtered).to_excel(writer, sheet_name="Filtered_Reviews", index=False)

    # タグ別比較シート
    for tag in available_tags:
        sheet_name = f"Tag_{tag[:25]}"
        tag_df = tag_ratio_long_df[tag_ratio_long_df["Tag"] == tag]
        pv = tag_df.pivot(index="Model", columns="Sentiment", values="Ratio(%)")
        reviews_map = tag_df.groupby("Model")["Reviews"].sum()
        pv["Reviews"] = pv.index.map(reviews_map).astype(int)
        pv = pv[SENTIMENTS + ["Reviews"]]
        _safe_df(pv).to_excel(writer, sheet_name=sheet_name)

    # 個別モデル集計シート
    for fname, df in file_dfs.items():
        sheet_name = f"{model_names[fname][:28]}_集計"
        _safe_df(build_tag_summary(df)).to_excel(writer, sheet_name=sheet_name)

buf.seek(0)
wb = load_workbook(buf)

# ---------------------------------------------------------------------------
# 画像貼付けユーティリティ
# ---------------------------------------------------------------------------
def _add_image(ws, img_data: bytes, cell: str, width: int, height: int):
    img = XLImage(img_data)
    img._id = None  # 重複 rId 回避
    img._name = f"Pic_{uuid4().hex[:8]}"
    img.width, img.height = width, height
    ws.add_image(img, cell)


# 個別モデルのヒストグラム画像
for fname, df in file_dfs.items():
    sheet = f"{model_names[fname][:28]}_集計"
    if sheet not in wb.sheetnames:
        continue
    ws = wb[sheet]
    fig, ax = plt.subplots()
    build_tag_summary(df).plot(kind="bar", stacked=True, ax=ax)
    ax.set_xlabel("Tag")
    ax.set_ylabel("Count")
    ax.legend(title="Sentiment")
    img_data = io.BytesIO()
    fig.savefig(img_data, format="png", bbox_inches="tight")
    plt.close(fig)
    img_data.seek(0)
    _add_image(ws, img_data, "H2", 480, 320)

# タグ別比較グラフ
for tag in available_tags:
    sheet = f"Tag_{tag[:25]}"
    if sheet not in wb.sheetnames:
        continue
    ws = wb[sheet]
    tag_df = tag_ratio_long_df[tag_ratio_long_df["Tag"] == tag]
    pv = tag_df.pivot(index="Model", columns="Sentiment", values="Ratio(%)")
    reviews_map = tag_df.groupby("Model")["Reviews"].sum()
    pv["Reviews"] = pv.index.map(reviews_map).astype(int)
    pv = pv[SENTIMENTS + ["Reviews"]]

    fig, ax = plt.subplots(figsize=(9, 7))
    pv[SENTIMENTS].plot(kind="bar", stacked=True, ax=ax)
    ax.set_ylabel("Percentage (%)")
    ax.set_title(tag)
    ax.legend(title="Sentiment", bbox_to_anchor=(1.05, 1), loc="upper left")
    for rect, total in zip(ax.containers[0], pv["Reviews"]):
        x_center = rect.get_x() + rect.get_width() / 2
        ax.text(x_center, 101, f"{int(total)}", ha="center", va="bottom", fontsize=8)
    ax.set_ylim(0, 105)
    fig.tight_layout()

    img_data = io.BytesIO()
    fig.savefig(img_data, format="png", bbox_inches="tight")
    plt.close(fig)
    img_data.seek(0)
    _add_image(ws, img_data, "H2", 800, 600)

# 保存
out = io.BytesIO()
wb.save(out)
out.seek(0)

st.download_button(
    "解析結果を Excel でダウンロード",
    data=out,
    file_name=f"bestbuy_review_analysis_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.success("集計完了!")
